import openpyxl
import re
import os
import json

excel_path = "/Users/vichy/Sites/temple2/excel/ฐานข้อมูลพระสังฆาธิการ_คณะสงฆ์ภาค2_อยุธยา_สระบุรี_อ่างทอง_2569_เบอร์โทรเลขล้วน.xlsx"

def clean_sql(val):
    if val is None:
        return ""
    return str(val).strip().replace("'", "''")

def clean_text(val):
    if val is None or str(val) == 'nan':
        return ""
    return str(val).strip()

def guess_pali_education(title, fullname, pos_sangha, pali_grade):
    if pali_grade:
        if "ป.ธ." in pali_grade:
            num = re.search(r'\d+', pali_grade)
            if num:
                return f"เปรียญธรรม {num.group(0)} ประโยค"
            return pali_grade
        return pali_grade
        
    pali = "ไม่มี"
    if title.startswith("พระมหา") or fullname.startswith("พระมหา"):
        pali = "เปรียญธรรม 3 ประโยค"
        match = re.search(r'ป\.ธ\.(\d)', pos_sangha)
        if match:
            pali = f"เปรียญธรรม {match.group(1)} ประโยค"
        else:
            match2 = re.search(r'ป\.ธ\.(\d)', title)
            if match2:
                pali = f"เปรียญธรรม {match2.group(1)} ประโยค"
            else:
                match3 = re.search(r'ป\.ธ\.(\d)', fullname)
                if match3:
                    pali = f"เปรียญธรรม {match3.group(1)} ประโยค"
    return pali

monks_list = []
temples_map = {} # (province, name) -> {name, type, district, subdistrict, province, abbot}

print("Loading Excel file...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['ข้อมูลรวม']

for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    code = row[0]
    if not code or str(code) == 'nan':
        continue

    person_code = clean_text(row[0])
    title = clean_text(row[1])
    first_name = clean_text(row[2])
    chaya = clean_text(row[3])
    pali_grade = clean_text(row[4])
    fullname = clean_text(row[5])
    temple = clean_text(row[6])
    subdistrict = clean_text(row[7])
    district = clean_text(row[8])
    province = clean_text(row[9])
    pos_temple = clean_text(row[10])
    pos_sangha = clean_text(row[11])
    phone_primary = clean_text(row[12])
    phone_secondary = clean_text(row[13])
    data_source = clean_text(row[14])

    pali = guess_pali_education(title, fullname, pos_sangha, pali_grade)
    
    is_abbot = any(x in pos_temple for x in ["เจ้าอาวาส", "จล.", "จร.", "รก.จร.", "รก.จล."])
    temple_type = "พระอารามหลวง" if ("จล." in pos_temple or "วิหาร" in temple or "วรวิหาร" in temple) else "วัดราษฎร์"

    if temple and province and province != "กรุงเทพมหานคร":
        key = (province, temple)
        if key not in temples_map:
            temples_map[key] = {
                'name': temple,
                'type': temple_type,
                'district': district,
                'subdistrict': subdistrict if subdistrict != '-' else '',
                'province': province,
                'abbot': fullname if is_abbot else ""
            }
        elif is_abbot and not temples_map[key]['abbot']:
            temples_map[key]['abbot'] = fullname

    region = "ภาค 2"

    monks_list.append({
        'personCode': person_code,
        'title': title,
        'firstName': first_name if first_name else title,
        'lastName': '',
        'chaya': chaya,
        'nickname': fullname,
        'phone': phone_primary,
        'phoneSecondary': phone_secondary,
        'residingTemple': temple,
        'affiliatedTemple': temple,
        'subdistrict': subdistrict if subdistrict != '-' else '',
        'district': district,
        'province': province,
        'region': region,
        'templePosition': pos_temple,
        'sanghaPosition': pos_sangha,
        'faction': 'มหานิกาย',
        'dhammaEducation': 'นักธรรมเอก',
        'paliEducation': pali,
        'paliGrade': pali_grade,
        'zipCode': '',
        'remarks': f"รหัสบุคคล: {person_code} | ที่มา: {data_source}",
        'dataSource': data_source,
        'upajjhayaStatus': 'ไม่มี',
        'upajjhayaCode': '',
        'otherPosition': '',
        'rajathinnanam': '',
        'rankClass': ''
    })

print(f"Parsed {len(monks_list)} monks and {len(temples_map)} temples.")

# Check if Saraburi Governor and Deputies exist, if not, append them
srb_gov_exists = any(m['province'] == 'สระบุรี' and 'จจ.สระบุรี' in m['sanghaPosition'] for m in monks_list)
if not srb_gov_exists:
    print("Appending Saraburi Provincial Hierarchy monks...")
    saraburi_governor = {
        'personCode': 'SRB-GOV',
        'title': 'พระราชวชิรมงคลวิสิฐ',
        'firstName': 'พระราชวชิรมงคลวิสิฐ',
        'lastName': '',
        'chaya': 'สิริวฑฺฒโน',
        'nickname': 'พระราชวชิรมงคลวิสิฐ สิริวฑฺฒโน',
        'phone': '036303112',
        'phoneSecondary': '',
        'residingTemple': 'มงคลชัยพัฒนา',
        'affiliatedTemple': 'มงคลชัยพัฒนา',
        'subdistrict': 'ห้วยบง',
        'district': 'เฉลิมพระเกียรติ',
        'province': 'สระบุรี',
        'region': 'ภาค 2',
        'templePosition': 'จล.',
        'sanghaPosition': 'จจ.สระบุรี',
        'faction': 'มหานิกาย',
        'dhammaEducation': 'นักธรรมเอก',
        'paliEducation': 'ไม่มี',
        'paliGrade': '',
        'zipCode': '18120',
        'remarks': 'เจ้าคณะจังหวัดสระบุรี',
        'dataSource': 'ข้อมูลเพิ่มเติมผู้บริหารจังหวัดสระบุรี',
        'upajjhayaStatus': 'ไม่มี',
        'upajjhayaCode': '',
        'otherPosition': '',
        'rajathinnanam': '',
        'rankClass': 'พระราชาคณะชั้นราช'
    }
    saraburi_deputy1 = {
        'personCode': 'SRB-DEP1',
        'title': 'พระครูศรีวรกิจจารักษ์',
        'firstName': 'พระครูศรีวรกิจจารักษ์',
        'lastName': '',
        'chaya': 'ถิรนาโถ',
        'nickname': 'พระครูศรีวรกิจจารักษ์ ถิรนาโถ',
        'phone': '036267111',
        'phoneSecondary': '',
        'residingTemple': 'พระพุทธบาท',
        'affiliatedTemple': 'พระพุทธบาท',
        'subdistrict': 'พระพุทธบาท',
        'district': 'พระพุทธบาท',
        'province': 'สระบุรี',
        'region': 'ภาค 2',
        'templePosition': 'ผู้ช่วยเจ้าอาวาส',
        'sanghaPosition': 'รจจ.สระบุรี',
        'faction': 'มหานิกาย',
        'dhammaEducation': 'นักธรรมเอก',
        'paliEducation': 'เปรียญธรรม ๖ ประโยค',
        'paliGrade': 'ป.ธ.6',
        'zipCode': '18120',
        'remarks': 'รองเจ้าคณะจังหวัดสระบุรี รูปที่ 1',
        'dataSource': 'ข้อมูลเพิ่มเติมผู้บริหารจังหวัดสระบุรี',
        'upajjhayaStatus': 'ไม่มี',
        'upajjhayaCode': '',
        'otherPosition': '',
        'rajathinnanam': '',
        'rankClass': 'พระครูสัญญาบัตร'
    }
    saraburi_deputy2 = {
        'personCode': 'SRB-DEP2',
        'title': 'พระมหาสมหมาย',
        'firstName': 'พระมหาสมหมาย',
        'lastName': '',
        'chaya': 'ธมฺมเสวี',
        'nickname': 'พระมหาสมหมาย ธมฺมเสวี ป.ธ.9',
        'phone': '036211234',
        'phoneSecondary': '',
        'residingTemple': 'ศรีบุรีรตนาราม',
        'affiliatedTemple': 'ศรีบุรีรตนาราม',
        'subdistrict': 'ปากเพรียว',
        'district': 'เมืองสระบุรี',
        'province': 'สระบุรี',
        'region': 'ภาค 2',
        'templePosition': 'เจ้าอาวาส',
        'sanghaPosition': 'รจจ.สระบุรี',
        'faction': 'มหานิกาย',
        'dhammaEducation': 'นักธรรมเอก',
        'paliEducation': 'เปรียญธรรม ๙ ประโยค',
        'paliGrade': 'ป.ธ.9',
        'zipCode': '18000',
        'remarks': 'รองเจ้าคณะจังหวัดสระบุรี รูปที่ 2',
        'dataSource': 'ข้อมูลเพิ่มเติมผู้บริหารจังหวัดสระบุรี',
        'upajjhayaStatus': 'ไม่มี',
        'upajjhayaCode': '',
        'otherPosition': '',
        'rajathinnanam': '',
        'rankClass': 'พระราชาคณะ'
    }
    monks_list.append(saraburi_governor)
    monks_list.append(saraburi_deputy1)
    monks_list.append(saraburi_deputy2)

# Generate import_data.sql
print("Generating import_data.sql...")
sql_statements = []
sql_statements.append("SET NAMES utf8mb4;")
sql_statements.append("SET CHARACTER SET utf8mb4;")
sql_statements.append("USE `temple2`;")
sql_statements.append("TRUNCATE TABLE `monks`;")
sql_statements.append("TRUNCATE TABLE `temples`;")

# Insert Monks SQL
for idx, monk in enumerate(monks_list, 1):
    sql = f"""INSERT INTO `monks` (
      `id`, `person_code`, `title`, `first_name`, `last_name`, `chaya`, `nickname`,
      `phone`, `phone_secondary`, `residing_temple`, `affiliated_temple`, 
      `subdistrict`, `district`, `province`, `region`, 
      `temple_position`, `sangha_position`, 
      `faction`, `dhamma_education`, `pali_education`, `pali_grade`,
      `zip_code`, `remarks`, `data_source`, `upajjhaya_status`, `upajjhaya_code`, `other_position`, `rajathinnanam`, `rank_class`
    ) VALUES (
      {idx}, '{clean_sql(monk['personCode'])}', '{clean_sql(monk['title'])}', '{clean_sql(monk['firstName'])}', '{clean_sql(monk['lastName'])}', '{clean_sql(monk['chaya'])}', '{clean_sql(monk['nickname'])}',
      '{clean_sql(monk['phone'])}', '{clean_sql(monk['phoneSecondary'])}', '{clean_sql(monk['residingTemple'])}', '{clean_sql(monk['affiliatedTemple'])}', 
      '{clean_sql(monk['subdistrict'])}', '{clean_sql(monk['district'])}', '{clean_sql(monk['province'])}', '{clean_sql(monk['region'])}', 
      '{clean_sql(monk['templePosition'])}', '{clean_sql(monk['sanghaPosition'])}', 
      '{clean_sql(monk['faction'])}', '{clean_sql(monk['dhammaEducation'])}', '{clean_sql(monk['paliEducation'])}', '{clean_sql(monk['paliGrade'])}',
      '{clean_sql(monk['zipCode'])}', '{clean_sql(monk['remarks'])}', '{clean_sql(monk['dataSource'])}', '{clean_sql(monk['upajjhayaStatus'])}', '{clean_sql(monk['upajjhayaCode'])}',
      '{clean_sql(monk['otherPosition'])}', '{clean_sql(monk['rajathinnanam'])}', '{clean_sql(monk['rankClass'])}'
    );"""
    sql_statements.append(sql)

# Insert Temples SQL
for idx, ((prov, temp_name), info) in enumerate(temples_map.items(), 1):
    abbot = info['abbot'] if info['abbot'] else "ไม่มีข้อมูลเจ้าอาวาส"
    sql = f"""INSERT INTO `temples` (
      `id`, `name`, `type`, `district`, `subdistrict`, `province`, `abbot`
    ) VALUES (
      {idx}, '{clean_sql(info['name'])}', '{clean_sql(info['type'])}', '{clean_sql(info['district'])}', '{clean_sql(info['subdistrict'])}', '{clean_sql(info['province'])}', '{clean_sql(abbot)}'
    );"""
    sql_statements.append(sql)

# Insert Events SQL
sql_statements.append("TRUNCATE TABLE `events`;")
events = [
    # Region 2 events
    ('การประชุมคณะสงฆ์ภาค 2 ประจำปี 2569', 'วันศุกร์ที่ 15 พฤษภาคม 2569', 'meeting', 'ภาค 2', 'ประชุมพระสังฆาธิการระดับเจ้าคณะภาค รองเจ้าคณะภาค เจ้าคณะจังหวัด และเลขาฯ คณะสงฆ์ภาค 2 ณ วัดพนัญเชิงวรวิหาร'),
    ('โครงการฝึกอบรมพระสังฆาธิการ คณะสงฆ์ภาค 2', 'วันจันทร์ที่ 22 มิถุนายน 2569', 'training', 'ภาค 2', 'การอบรมเชิงปฏิบัติการพัฒนาศักยภาพการบริหารจัดการวัดและทะเบียนสงฆ์ สำหรับเจ้าคณะอำเภอและเลขาฯ ในเขตภาค 2'),
    
    # Ayutthaya events
    ('วันพระ (ขึ้น ๘ ค่ำ เดือน ๘)', 'วันอาทิตย์ที่ 21 มิถุนายน 2569', 'holy-day', 'พระนครศรีอยุธยา', 'ทำบุญตักบาตร ฟังพระธรรมเทศนา ณ วัดใกล้บ้านในเขตพื้นที่'),
    ('วันอาสาฬหบูชา (ขึ้น ๑๕ ค่ำ เดือน ๘)', 'วันอาทิตย์ที่ 28 มิถุนายน 2569', 'holy-day', 'พระนครศรีอยุธยา', 'พิธีเวียนเทียน ถวายเทียนพรรษา และเทศนาธรรมใหญ่ประจำปี ณ วัดพนัญเชิงวรวิหาร และวัดต่างๆ ทุกตำบล'),
    ('การประชุมคณะสงฆ์จังหวัดพระนครศรีอยุธยา ประจำเดือน', 'วันพุธที่ 1 กรกฎาคม 2569', 'meeting', 'พระนครศรีอยุธยา', 'ประชุมพระสังฆาธิการระดับเจ้าคณะตำบลและเจ้าคณะอำเภอ ณ ศาลาการเปรียญวัดพนัญเชิงวรวิหาร เวลา 13.00 น.'),
    
    # Ang Thong events
    ('วันพระ (ขึ้น ๘ ค่ำ เดือน ๘)', 'วันอาทิตย์ที่ 21 มิถุนายน 2569', 'holy-day', 'อ่างทอง', 'ทำบุญตักบาตร ถวายภัตตาหารเช้า ณ วัดต่างๆ ในอำเภอเมืองอ่างทอง'),
    ('วันอาสาฬหบูชา (ขึ้น ๑๕ ค่ำ เดือน ๘)', 'วันอาทิตย์ที่ 28 มิถุนายน 2569', 'holy-day', 'อ่างทอง', 'พิธีเวียนเทียน ถวายเทียนพรรษา ณ วัดต้นสน และวัดอ่างทองวรวิหาร'),
    ('การประชุมคณะสงฆ์จังหวัดอ่างทอง ประจำเดือน', 'วันพฤหัสบดีที่ 2 กรกฎาคม 2569', 'meeting', 'อ่างทอง', 'ประชุมพระสังฆาธิการระดับเจ้าคณะตำบลและอำเภอ ณ ศาลาการเปรียญวัดอ่างทองวรวิหาร เวลา 13.00 น.'),
    
    # Saraburi events
    ('วันพระ (ขึ้น ๘ ค่ำ เดือน ๘)', 'วันอาทิตย์ที่ 21 มิถุนายน 2569', 'holy-day', 'สระบุรี', 'ทำบุญตักบาตร รักษาศีลฟังธรรม ณ วัดใกล้บ้านในเขตพื้นที่จังหวัดสระบุรี'),
    ('วันอาสาฬหบูชา (ขึ้น ๑๕ ค่ำ เดือน ๘)', 'วันอาทิตย์ที่ 28 มิถุนายน 2569', 'holy-day', 'สระบุรี', 'พิธีเวียนเทียนและเทศนาธรรมใหญ่ประจำปี ณ วัดพระพุทธบาทราชวรมหาวิหาร'),
    ('การประชุมคณะสงฆ์จังหวัดสระบุรี ประจำเดือน', 'วันศุกร์ที่ 3 กรกฎาคม 2569', 'meeting', 'สระบุรี', 'ประชุมพระสังฆาธิการระดับเจ้าคณะตำบลและอำเภอ ณ วัดมงคลชัยพัฒนา เวลา 13.00 น.')
]

for idx, (title, date, etype, prov, desc) in enumerate(events, 1):
    sql = f"""INSERT INTO `events` (
      `id`, `title`, `date`, `type`, `province`, `description`
    ) VALUES (
      {idx}, '{clean_sql(title)}', '{clean_sql(date)}', '{clean_sql(etype)}', '{clean_sql(prov)}', '{clean_sql(desc)}'
    );"""
    sql_statements.append(sql)

with open("/Users/vichy/Sites/temple2/import_data.sql", "w", encoding="utf-8") as f:
    f.write("\n".join(sql_statements))
print("Successfully generated import_data.sql.")

# Generate data.json
print("Generating data.json...")
json_monks = []
for monk in monks_list:
    json_monks.append({
        'personCode': monk['personCode'],
        'title': monk['title'],
        'firstName': monk['firstName'],
        'lastName': monk['lastName'],
        'chaya': monk['chaya'],
        'nickname': monk['nickname'],
        'idCard': '',
        'birthDate': '',
        'phone': monk['phone'],
        'phoneSecondary': monk['phoneSecondary'],
        'lineId': '',
        'ordinationDate': '',
        'upajjhaya': '',
        'vassa': 0,
        'age': 0,
        'residingTemple': monk['residingTemple'],
        'affiliatedTemple': monk['affiliatedTemple'],
        'subdistrict': monk['subdistrict'],
        'district': monk['district'],
        'province': monk['province'],
        'region': monk['region'],
        'templePosition': monk['templePosition'],
        'sanghaPosition': monk['sanghaPosition'],
        'upajjhayaStatus': monk['upajjhayaStatus'],
        'upajjhayaCode': monk['upajjhayaCode'],
        'otherPosition': monk['otherPosition'],
        'rajathinnanam': monk['rajathinnanam'],
        'rankClass': monk['rankClass'],
        'faction': monk['faction'],
        'education': '',
        'dhammaEducation': monk['dhammaEducation'],
        'paliEducation': monk['paliEducation'],
        'paliGrade': monk['paliGrade'],
        'zipCode': monk['zipCode'],
        'remarks': monk['remarks'],
        'dataSource': monk['dataSource']
    })

json_temples = []
for idx, ((prov, temp_name), info) in enumerate(temples_map.items(), 1):
    json_temples.append({
        'id': idx,
        'name': info['name'],
        'type': info['type'],
        'district': info['district'],
        'subdistrict': info['subdistrict'],
        'province': info['province'],
        'abbot': info['abbot'] if info['abbot'] else "ไม่มีข้อมูลเจ้าอาวาส"
    })

json_events = []
for idx, (title, date, etype, prov, desc) in enumerate(events, 1):
    json_events.append({
        'id': idx,
        'title': title,
        'date': date,
        'type': etype,
        'province': prov,
        'description': desc
    })

data_json_content = {
    'status': 'success',
    'monks': json_monks,
    'temples': json_temples,
    'events': json_events
}

with open("/Users/vichy/Sites/temple2/data.json", "w", encoding="utf-8") as f:
    json.dump(data_json_content, f, ensure_ascii=False, indent=2)
print("Successfully generated data.json.")

# Generate data.js
print("Generating data.js...")
js_content = f"window.SANGHA_DATA_FALLBACK = {json.dumps(data_json_content, ensure_ascii=False, indent=2)};"
with open("/Users/vichy/Sites/temple2/data.js", "w", encoding="utf-8") as f:
    f.write(js_content)
print("Successfully generated data.js.")
